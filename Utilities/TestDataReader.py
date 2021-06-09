import openpyxl


def data_generator():
    workbook = openpyxl.load_workbook("./TestData/LoginTestData.xlsx")
    worksheet = workbook['Sheet1']
    rows = worksheet.max_row
    final_list = []
    inner_list = []
    for row in range(2, rows+1):
        inner_list = []
        inner_list = [worksheet.cell(row, 1).value, worksheet.cell(row, 2).value]
        final_list.append(inner_list)
    return final_list
